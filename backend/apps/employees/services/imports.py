import csv
import io
from dataclasses import dataclass
from datetime import date, datetime
from typing import Any

from django.core.exceptions import ValidationError
from django.db import transaction

from apps.authentication.models import UserRole
from apps.core.utils.text import normalize_code
from apps.employees.models import (
    EmployeeProfile,
    EmploymentStatus,
    Gender,
)
from apps.organization.models import (
    Department,
    Designation,
    Site,
)


EMPLOYEE_IMPORT_COLUMNS = [
    "employee_id",
    "first_name",
    "last_name",
    "email",
    "mobile",
    "gender",
    "date_of_joining",
    "employment_status",
    "site_code",
    "department_code",
    "designation_code",
    "reporting_manager_employee_id",
    "role",
    "is_active",
    "erp_user_id",
    "last_working_date",
]

REQUIRED_IMPORT_COLUMNS = [
    "first_name",
]

DATE_FORMATS = [
    "%Y-%m-%d",
    "%d-%m-%Y",
    "%d/%m/%Y",
]


@dataclass(frozen=True)
class ImportRowResult:
    row_number: int
    row: dict[str, Any]
    normalized: dict[str, Any]
    errors: list[str]

    @property
    def is_valid(self) -> bool:
        return not self.errors


def build_csv_template() -> str:
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(EMPLOYEE_IMPORT_COLUMNS)
    writer.writerow(
        [
            "",
            "Asha",
            "Sharma",
            "asha.sharma@jnl.com",
            "9876543210",
            Gender.FEMALE,
            "2026-07-21",
            EmploymentStatus.CONFIRMED,
            "BKN",
            "FIN",
            "HOD",
            "",
            UserRole.USER,
            "true",
            "ERP001",
            "",
        ]
    )

    return output.getvalue()


def build_xlsx_template() -> bytes:
    try:
        from openpyxl import Workbook
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError(
            "Excel template support requires openpyxl."
        ) from exc

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Employees"
    worksheet.append(EMPLOYEE_IMPORT_COLUMNS)
    worksheet.append(
        [
            "",
            "Asha",
            "Sharma",
            "asha.sharma@jnl.com",
            "9876543210",
            Gender.FEMALE,
            "2026-07-21",
            EmploymentStatus.CONFIRMED,
            "BKN",
            "FIN",
            "HOD",
            "",
            UserRole.USER,
            "true",
            "ERP001",
            "",
        ]
    )

    output = io.BytesIO()
    workbook.save(output)

    return output.getvalue()


def parse_employee_import_file(uploaded_file):
    filename = uploaded_file.name.lower()

    if filename.endswith(".csv"):
        return _parse_csv(uploaded_file)

    if filename.endswith((".xlsx", ".xlsm")):
        return _parse_xlsx(uploaded_file)

    raise ValidationError(
        "Upload a CSV or XLSX employee import file."
    )


def preview_employee_import(uploaded_file):
    rows = parse_employee_import_file(uploaded_file)
    results = validate_import_rows(rows)
    valid_rows = [
        result for result in results if result.is_valid
    ]
    failed_rows = [
        result for result in results if not result.is_valid
    ]

    return {
        "summary": {
            "total_rows": len(results),
            "valid_rows": len(valid_rows),
            "failed_rows": len(failed_rows),
        },
        "rows": [
            _serialize_result(result)
            for result in results
        ],
        "failed_rows": [
            _serialize_result(result)
            for result in failed_rows
        ],
    }


@transaction.atomic
def import_employee_rows(uploaded_file):
    rows = parse_employee_import_file(uploaded_file)
    results = validate_import_rows(rows)
    created_profiles = []
    failed_rows = []

    for result in results:
        if not result.is_valid:
            failed_rows.append(result)
            continue

        try:
            created_profiles.append(
                EmployeeProfile.objects.create(
                    **result.normalized
                )
            )
        except ValidationError as exc:
            failed_rows.append(
                ImportRowResult(
                    row_number=result.row_number,
                    row=result.row,
                    normalized=result.normalized,
                    errors=[
                        str(message)
                        for messages in exc.message_dict.values()
                        for message in messages
                    ],
                )
            )

    return {
        "summary": {
            "total_rows": len(results),
            "created_rows": len(created_profiles),
            "failed_rows": len(failed_rows),
        },
        "created_employee_ids": [
            profile.employee_id
            for profile in created_profiles
        ],
        "failed_rows": [
            _serialize_result(result)
            for result in failed_rows
        ],
    }


def validate_import_rows(rows):
    seen_employee_ids = set()
    duplicate_employee_ids = set()

    for row in rows:
        employee_id = _normalize_optional_code(
            row.get("employee_id")
        )
        if not employee_id:
            continue
        if employee_id in seen_employee_ids:
            duplicate_employee_ids.add(employee_id)
        seen_employee_ids.add(employee_id)

    existing_employee_ids = set(
        EmployeeProfile.objects.filter(
            employee_id__in=seen_employee_ids
        ).values_list("employee_id", flat=True)
    )

    site_map = {
        site.site_code: site
        for site in Site.objects.select_related(
            "company"
        )
    }
    department_map = {
        department.department_code: department
        for department in Department.objects.select_related(
            "company"
        )
    }
    designation_map = {
        designation.designation_code: designation
        for designation in Designation.objects.all()
    }
    manager_map = {
        profile.employee_id: profile
        for profile in EmployeeProfile.objects.all()
    }

    results = []

    for index, row in enumerate(rows, start=2):
        result = _validate_row(
            row=row,
            row_number=index,
            duplicate_employee_ids=duplicate_employee_ids,
            existing_employee_ids=existing_employee_ids,
            site_map=site_map,
            department_map=department_map,
            designation_map=designation_map,
            manager_map=manager_map,
        )
        results.append(result)

    return results


def _parse_csv(uploaded_file):
    uploaded_file.seek(0)
    content = uploaded_file.read()

    if isinstance(content, bytes):
        text = content.decode("utf-8-sig")
    else:
        text = content

    reader = csv.DictReader(io.StringIO(text))
    _validate_headers(reader.fieldnames or [])

    return [
        {
            column: (row.get(column) or "").strip()
            for column in EMPLOYEE_IMPORT_COLUMNS
        }
        for row in reader
    ]


def _parse_xlsx(uploaded_file):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError(
            "Excel import support requires openpyxl."
        ) from exc

    uploaded_file.seek(0)
    workbook = load_workbook(
        uploaded_file,
        data_only=True,
        read_only=True,
    )
    worksheet = workbook.active
    rows = list(worksheet.iter_rows(values_only=True))

    if not rows:
        raise ValidationError(
            "The import file does not contain a header row."
        )

    headers = [
        str(value or "").strip()
        for value in rows[0]
    ]
    _validate_headers(headers)

    parsed_rows = []
    for values in rows[1:]:
        parsed_rows.append(
            {
                column: _stringify_cell(
                    values[headers.index(column)]
                    if column in headers
                    and headers.index(column)
                    < len(values)
                    else ""
                )
                for column in EMPLOYEE_IMPORT_COLUMNS
            }
        )

    return parsed_rows


def _validate_headers(headers):
    missing_headers = [
        column
        for column in REQUIRED_IMPORT_COLUMNS
        if column not in headers
    ]

    if missing_headers:
        raise ValidationError(
            "Missing required import columns: "
            + ", ".join(missing_headers)
        )


def _validate_row(
    *,
    row,
    row_number,
    duplicate_employee_ids,
    existing_employee_ids,
    site_map,
    department_map,
    designation_map,
    manager_map,
):
    errors = []
    normalized = {}

    employee_id = _normalize_optional_code(
        row.get("employee_id")
    )
    first_name = (row.get("first_name") or "").strip()

    if employee_id in duplicate_employee_ids:
        errors.append(
            "Duplicate Employee ID in import file."
        )
    elif employee_id in existing_employee_ids:
        errors.append(
            "Employee ID already exists."
        )
    else:
        normalized["employee_id"] = employee_id

    if not first_name:
        errors.append("First name is required.")
    else:
        normalized["first_name"] = first_name

    normalized["last_name"] = (
        row.get("last_name") or ""
    ).strip()
    normalized["email"] = (
        row.get("email") or ""
    ).strip()
    normalized["mobile"] = (
        row.get("mobile") or ""
    ).strip()
    normalized["gender"] = _choice_value(
        row.get("gender"),
        Gender,
        Gender.NOT_SPECIFIED,
        "Gender",
        errors,
    )
    normalized["employment_status"] = _choice_value(
        row.get("employment_status"),
        EmploymentStatus,
        EmploymentStatus.CONFIRMED,
        "Employment status",
        errors,
    )
    normalized["role"] = _choice_value(
        row.get("role"),
        UserRole,
        UserRole.USER,
        "Role",
        errors,
    )
    normalized["date_of_joining"] = _parse_date(
        row.get("date_of_joining"),
        "Date of joining",
        errors,
    )
    normalized["last_working_date"] = _parse_date(
        row.get("last_working_date"),
        "Last working date",
        errors,
    )
    normalized["is_active"] = _parse_bool(
        row.get("is_active"),
        default=True,
    )
    normalized["erp_user_id"] = _normalize_optional_code(
        row.get("erp_user_id")
    )

    site = _mapped_object(
        row.get("site_code"),
        site_map,
        "Site",
        errors,
    )
    department = _mapped_object(
        row.get("department_code"),
        department_map,
        "Department",
        errors,
    )
    designation = _mapped_object(
        row.get("designation_code"),
        designation_map,
        "Designation",
        errors,
    )
    reporting_manager = _mapped_object(
        row.get("reporting_manager_employee_id"),
        manager_map,
        "Reporting manager",
        errors,
    )

    normalized["site"] = site
    normalized["department"] = department
    normalized["designation"] = designation
    normalized["reporting_manager"] = reporting_manager

    if (
        site
        and department
        and site.company_id != department.company_id
    ):
        errors.append(
            "Department must belong to the same company as the site."
        )

    if (
        employee_id
        and reporting_manager
        and reporting_manager.employee_id == employee_id
    ):
        errors.append(
            "Employee cannot report to themselves."
        )

    return ImportRowResult(
        row_number=row_number,
        row=row,
        normalized=normalized,
        errors=errors,
    )


def _serialize_result(result):
    normalized = {
        key: (
            value.id
            if hasattr(value, "id")
            else value.isoformat()
            if isinstance(value, date)
            else value
        )
        for key, value in result.normalized.items()
    }

    return {
        "row_number": result.row_number,
        "row": result.row,
        "normalized": normalized,
        "errors": result.errors,
        "is_valid": result.is_valid,
    }


def _stringify_cell(value):
    if value is None:
        return ""

    if isinstance(value, (date, datetime)):
        return value.strftime("%Y-%m-%d")

    return str(value).strip()


def _normalize_optional_code(value):
    if value is None or str(value).strip() == "":
        return ""

    return normalize_code(str(value))


def _parse_date(value, label, errors):
    if value is None or str(value).strip() == "":
        return None

    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, date):
        return value

    raw_value = str(value).strip()
    for date_format in DATE_FORMATS:
        try:
            return datetime.strptime(
                raw_value,
                date_format,
            ).date()
        except ValueError:
            continue

    errors.append(
        f"{label} must be a valid date."
    )
    return None


def _parse_bool(value, default):
    if value is None or str(value).strip() == "":
        return default

    return str(value).strip().lower() in {
        "1",
        "true",
        "yes",
        "y",
        "active",
    }


def _choice_value(
    value,
    choices_class,
    default,
    label,
    errors,
):
    if value is None or str(value).strip() == "":
        return default

    normalized_value = normalize_code(str(value))
    valid_values = {
        choice.value for choice in choices_class
    }

    if normalized_value not in valid_values:
        errors.append(
            f"{label} must be one of: "
            + ", ".join(sorted(valid_values))
        )
        return default

    return normalized_value


def _mapped_object(value, object_map, label, errors):
    code = _normalize_optional_code(value)

    if not code:
        return None

    mapped_value = object_map.get(code)

    if mapped_value is None:
        errors.append(
            f"{label} mapping not found for {code}."
        )

    return mapped_value
