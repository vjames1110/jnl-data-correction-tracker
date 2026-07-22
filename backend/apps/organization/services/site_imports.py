import csv
import io
from dataclasses import dataclass
from datetime import date, datetime
from typing import Any

from django.core.exceptions import ValidationError
from django.db import transaction

from apps.authentication.models import User
from apps.core.utils.text import normalize_code
from apps.organization.models import Company, Site


SITE_IMPORT_COLUMNS = [
    "company_code",
    "site_code",
    "site_name",
    "project_name",
    "state",
    "district",
    "address",
    "start_date",
    "end_date",
    "site_director_employee_id",
    "project_manager_employee_id",
    "cost_centre",
    "erp_site_code",
    "is_active",
]

REQUIRED_SITE_IMPORT_COLUMNS = [
    "company_code",
    "site_code",
    "site_name",
]

DATE_FORMATS = [
    "%Y-%m-%d",
    "%d-%m-%Y",
    "%d/%m/%Y",
]


@dataclass(frozen=True)
class SiteImportRowResult:
    row_number: int
    row: dict[str, Any]
    normalized: dict[str, Any]
    errors: list[str]

    @property
    def is_valid(self) -> bool:
        return not self.errors


def build_site_csv_template() -> str:
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(SITE_IMPORT_COLUMNS)
    writer.writerow(
        [
            "JNL",
            "BKN",
            "Bikaner Site",
            "Bikaner Road Project",
            "Rajasthan",
            "Bikaner",
            "",
            "2026-07-22",
            "",
            "JNLDIR00001",
            "JNLEMP00001",
            "CC001",
            "ERP_BKN",
            "true",
        ]
    )

    return output.getvalue()


def build_site_xlsx_template() -> bytes:
    try:
        from openpyxl import Workbook
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError(
            "Excel template support requires openpyxl."
        ) from exc

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Sites"
    worksheet.append(SITE_IMPORT_COLUMNS)
    worksheet.append(
        [
            "JNL",
            "BKN",
            "Bikaner Site",
            "Bikaner Road Project",
            "Rajasthan",
            "Bikaner",
            "",
            "2026-07-22",
            "",
            "JNLDIR00001",
            "JNLEMP00001",
            "CC001",
            "ERP_BKN",
            "true",
        ]
    )

    output = io.BytesIO()
    workbook.save(output)

    return output.getvalue()


def preview_site_import(uploaded_file):
    rows = parse_site_import_file(uploaded_file)
    results = validate_site_import_rows(rows)
    failed_rows = [
        result for result in results if not result.is_valid
    ]

    return {
        "summary": {
            "total_rows": len(results),
            "valid_rows": len(results) - len(failed_rows),
            "failed_rows": len(failed_rows),
        },
        "rows": [
            _serialize_result(result)
            for result in results
        ],
        "failed_rows": [
            _serialize_result(result)
            for result in failed_rows
        ],
    }


@transaction.atomic
def import_site_rows(uploaded_file):
    rows = parse_site_import_file(uploaded_file)
    results = validate_site_import_rows(rows)
    created_sites = []
    failed_rows = []

    for result in results:
        if not result.is_valid:
            failed_rows.append(result)
            continue

        try:
            created_sites.append(
                Site.objects.create(**result.normalized)
            )
        except ValidationError as exc:
            failed_rows.append(
                SiteImportRowResult(
                    row_number=result.row_number,
                    row=result.row,
                    normalized=result.normalized,
                    errors=[
                        str(message)
                        for messages in exc.message_dict.values()
                        for message in messages
                    ],
                )
            )

    return {
        "summary": {
            "total_rows": len(results),
            "created_rows": len(created_sites),
            "failed_rows": len(failed_rows),
        },
        "created_site_codes": [
            site.site_code for site in created_sites
        ],
        "failed_rows": [
            _serialize_result(result)
            for result in failed_rows
        ],
    }


def parse_site_import_file(uploaded_file):
    filename = uploaded_file.name.lower()

    if filename.endswith(".csv"):
        return _parse_csv(uploaded_file)

    if filename.endswith((".xlsx", ".xlsm")):
        return _parse_xlsx(uploaded_file)

    raise ValidationError(
        "Upload a CSV or XLSX site import file."
    )


def validate_site_import_rows(rows):
    company_map = {
        company.company_code: company
        for company in Company.objects.all()
    }
    user_map = {
        user.employee_id: user
        for user in User.objects.filter(is_active=True)
    }
    existing_site_keys = set(
        Site.objects.values_list(
            "company__company_code",
            "site_code",
        )
    )
    seen_site_keys = set()
    duplicate_site_keys = set()

    for row in rows:
        company_code = _normalize_optional_code(
            row.get("company_code")
        )
        site_code = _normalize_optional_code(
            row.get("site_code")
        )
        if not company_code or not site_code:
            continue
        key = (company_code, site_code)
        if key in seen_site_keys:
            duplicate_site_keys.add(key)
        seen_site_keys.add(key)

    return [
        _validate_row(
            row=row,
            row_number=index,
            company_map=company_map,
            user_map=user_map,
            existing_site_keys=existing_site_keys,
            duplicate_site_keys=duplicate_site_keys,
        )
        for index, row in enumerate(rows, start=2)
    ]


def _validate_row(
    *,
    row,
    row_number,
    company_map,
    user_map,
    existing_site_keys,
    duplicate_site_keys,
):
    errors = []
    normalized = {}
    company_code = _normalize_optional_code(
        row.get("company_code")
    )
    site_code = _normalize_optional_code(
        row.get("site_code")
    )
    site_name = (row.get("site_name") or "").strip()

    company = _mapped_object(
        company_code,
        company_map,
        "Company",
        errors,
    )

    if site_code:
        normalized["site_code"] = site_code
    else:
        errors.append("Site code is required.")

    if site_name:
        normalized["site_name"] = site_name
    else:
        errors.append("Site name is required.")

    if company:
        normalized["company"] = company
        key = (company.company_code, site_code)
        if key in duplicate_site_keys:
            errors.append(
                "Duplicate site code in import file."
            )
        elif key in existing_site_keys:
            errors.append(
                "Site code already exists for this company."
            )

    normalized["project_name"] = (
        row.get("project_name") or ""
    ).strip()
    normalized["state"] = (
        row.get("state") or ""
    ).strip()
    normalized["district"] = (
        row.get("district") or ""
    ).strip()
    normalized["address"] = (
        row.get("address") or ""
    ).strip()
    normalized["start_date"] = _parse_date(
        row.get("start_date"),
        "Start date",
        errors,
    )
    normalized["end_date"] = _parse_date(
        row.get("end_date"),
        "End date",
        errors,
    )
    normalized["site_director"] = _mapped_object(
        row.get("site_director_employee_id"),
        user_map,
        "Site director",
        errors,
    )
    normalized["site_hod"] = _mapped_object(
        row.get("project_manager_employee_id"),
        user_map,
        "Project manager",
        errors,
    )
    normalized["cost_centre"] = _normalize_optional_code(
        row.get("cost_centre")
    )
    normalized["erp_site_code"] = _normalize_optional_code(
        row.get("erp_site_code")
    )
    normalized["is_active"] = _parse_bool(
        row.get("is_active"),
        default=True,
    )

    return SiteImportRowResult(
        row_number=row_number,
        row=row,
        normalized=normalized,
        errors=errors,
    )


def _parse_csv(uploaded_file):
    uploaded_file.seek(0)
    content = uploaded_file.read()
    text = (
        content.decode("utf-8-sig")
        if isinstance(content, bytes)
        else content
    )
    reader = csv.DictReader(io.StringIO(text))
    _validate_headers(reader.fieldnames or [])

    return [
        {
            column: (row.get(column) or "").strip()
            for column in SITE_IMPORT_COLUMNS
        }
        for row in reader
    ]


def _parse_xlsx(uploaded_file):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError(
            "Excel import support requires openpyxl."
        ) from exc

    uploaded_file.seek(0)
    workbook = load_workbook(
        uploaded_file,
        data_only=True,
        read_only=True,
    )
    worksheet = workbook.active
    rows = list(worksheet.iter_rows(values_only=True))

    if not rows:
        raise ValidationError(
            "The import file does not contain a header row."
        )

    headers = [
        str(value or "").strip()
        for value in rows[0]
    ]
    _validate_headers(headers)

    parsed_rows = []
    for values in rows[1:]:
        parsed_rows.append(
            {
                column: _stringify_cell(
                    values[headers.index(column)]
                    if column in headers
                    and headers.index(column)
                    < len(values)
                    else ""
                )
                for column in SITE_IMPORT_COLUMNS
            }
        )

    return parsed_rows


def _validate_headers(headers):
    missing_headers = [
        column
        for column in REQUIRED_SITE_IMPORT_COLUMNS
        if column not in headers
    ]

    if missing_headers:
        raise ValidationError(
            "Missing required import columns: "
            + ", ".join(missing_headers)
        )


def _serialize_result(result):
    normalized = {
        key: (
            value.id
            if hasattr(value, "id")
            else value.isoformat()
            if isinstance(value, date)
            else value
        )
        for key, value in result.normalized.items()
    }

    return {
        "row_number": result.row_number,
        "row": result.row,
        "normalized": normalized,
        "errors": result.errors,
        "is_valid": result.is_valid,
    }


def _stringify_cell(value):
    if value is None:
        return ""

    if isinstance(value, (date, datetime)):
        return value.strftime("%Y-%m-%d")

    return str(value).strip()


def _normalize_optional_code(value):
    if value is None or str(value).strip() == "":
        return ""

    return normalize_code(str(value))


def _mapped_object(value, mapping, label, errors):
    normalized_value = _normalize_optional_code(value)
    if not normalized_value:
        if label == "Company":
            errors.append("Company code is required.")
        return None

    matched = mapping.get(normalized_value)
    if matched is None:
        errors.append(f"{label} mapping not found.")

    return matched


def _parse_date(value, label, errors):
    value = (value or "").strip()
    if not value:
        return None

    for date_format in DATE_FORMATS:
        try:
            return datetime.strptime(
                value,
                date_format,
            ).date()
        except ValueError:
            continue

    errors.append(
        f"{label} must be a valid date."
    )
    return None


def _parse_bool(value, default=True):
    if value is None or str(value).strip() == "":
        return default

    return str(value).strip().lower() in {
        "1",
        "true",
        "yes",
        "y",
        "active",
    }
